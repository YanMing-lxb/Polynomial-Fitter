import argparse
import os
import sys

import numpy as np
import rich_argparse
from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.prompt import IntPrompt, Prompt
from rich.table import Table
from scipy.optimize import curve_fit

from plot_show import plot_fits_pygal_combined
from utils import polynomial_fit
from version import script_name

console = Console()


def read_excel_data(file_path: str) -> dict[str, np.ndarray]:
    """ä»Excelæ–‡ä»¶ä¸­è¯»å–æ•°æ®ï¼Œæ”¯æŒåŒå±‚è¡¨å¤´ç»“æ„

    è¯¥å‡½æ•°è¯»å–æŒ‡å®šè·¯å¾„çš„Excelæ–‡ä»¶ï¼Œå°†å‰ä¸¤è¡Œä½œä¸ºåŒå±‚è¡¨å¤´å¤„ç†ï¼Œ
    ä»ç¬¬ä¸‰è¡Œå¼€å§‹ä½œä¸ºæ•°æ®è¡Œè¯»å–ï¼Œå¹¶å°†æ•°æ®ç»„ç»‡æˆå­—å…¸ç»“æ„ï¼Œ
    å…¶ä¸­é”®ä¸ºåŒå±‚è¡¨å¤´çš„å…ƒç»„ï¼Œå€¼ä¸ºå¯¹åº”çš„NumPyæ•°ç»„ã€‚

    Parameters
    ----------
    file_path : str
        Excelæ–‡ä»¶çš„è·¯å¾„ï¼Œæ”¯æŒ.xlsxæ ¼å¼

    Returns
    -------
    dict[tuple[str, str], np.ndarray]
        ä»¥åŒå±‚è¡¨å¤´å…ƒç»„ä¸ºé”®ï¼Œå¯¹åº”åˆ—æ•°æ®çš„NumPyæ•°ç»„ä¸ºå€¼çš„å­—å…¸
    """
    if not isinstance(file_path, str) or not file_path.endswith(".xlsx"):
        raise ValueError("file_path å¿…é¡»æ˜¯éç©ºçš„ .xlsx æ–‡ä»¶è·¯å¾„")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"æ–‡ä»¶æœªæ‰¾åˆ°: {file_path}")

    wb = None
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("Excel æ–‡ä»¶è‡³å°‘éœ€è¦ä¸¤è¡Œä½œä¸ºåŒå±‚è¡¨å¤´")

        # è¯»å–å‰ä¸¤è¡Œä½œä¸ºåŒå±‚ header
        header_row1 = rows[0]
        header_row2 = rows[1]

        combined_headers = [tuple(pair) for pair in zip(header_row1, header_row2)]

        # åˆå§‹åŒ–æ¯åˆ—çš„æ•°æ®åˆ—è¡¨
        data_dict = {header: [] for header in combined_headers}

        # ä»ç¬¬ä¸‰è¡Œå¼€å§‹è¯»å–æ•°æ®
        for row in rows[2:]:
            for i, value in enumerate(row):
                if i < len(combined_headers):
                    data_dict[combined_headers[i]].append(value)

        # è½¬æ¢ä¸º NumPy æ•°ç»„å¹¶è¿”å›
        return {k: np.array(v) for k, v in data_dict.items()}

    except Exception as e:
        raise RuntimeError(f"è¯»å– Excel æ•°æ®æ—¶å‡ºé”™: {e}") from e
    finally:
        if wb is not None:
            wb.close()


def fit_polynomial(
    data: dict[str, np.ndarray], max_degree: int = 9, threshold: float = 0.9999
) -> dict[str, dict[str, int | float | np.ndarray]]:
    """
    å¯¹ç»™å®šæ•°æ®è¿›è¡Œå¤šé¡¹å¼æ‹Ÿåˆï¼Œå¯»æ‰¾æœ€ä¼˜æ‹Ÿåˆåº¦æ»¡è¶³é˜ˆå€¼è¦æ±‚çš„æœ€ä½é˜¶å¤šé¡¹å¼æ¨¡å‹ã€‚

    å‚æ•°:
    data : dict[str, np.ndarray]
        è¾“å…¥æ•°æ®å­—å…¸ï¼Œå…¶ä¸­ç¬¬ä¸€ä¸ªé”®å¯¹åº”çš„å€¼ä¸ºè‡ªå˜é‡ x çš„æ•°æ®ï¼Œ
        å…¶ä½™é”®å¯¹åº”å› å˜é‡ y çš„æ•°æ®ã€‚æ‰€æœ‰å€¼åº”ä¸º numpy æ•°ç»„ã€‚
    max_degree : int, optional
        æ‹Ÿåˆå¤šé¡¹å¼çš„æœ€é«˜é˜¶æ•°ï¼Œé»˜è®¤ä¸º 9ã€‚
    threshold : float, optional
        åˆ¤å®šæ‹Ÿåˆä¼˜åº¦æ˜¯å¦åˆæ ¼çš„ RÂ² å’Œè°ƒæ•´ RÂ² é˜ˆå€¼ï¼Œé»˜è®¤ä¸º 0.9999ã€‚

    è¿”å›:
    dict[str, dict[str, int | float | np.ndarray]]
        æ¯ä¸ªå› å˜é‡çš„æ‹Ÿåˆç»“æœå­—å…¸ï¼š
        - è‹¥æˆåŠŸæ‹Ÿåˆï¼Œåˆ™åŒ…å« "degree"ï¼ˆé˜¶æ•°ï¼‰ã€"coefficients"ï¼ˆç³»æ•°ï¼‰ã€
          "r_squared"ï¼ˆRÂ²ï¼‰å’Œ "adj_r_squared"ï¼ˆè°ƒæ•´ RÂ²ï¼‰ï¼›
        - è‹¥æ‹Ÿåˆå¤±è´¥æˆ–æ— æ³•æ”¶æ•›ï¼Œåˆ™è¿”å›é”™è¯¯ä¿¡æ¯å­—ç¬¦ä¸²ã€‚
    """
    results = {}

    # æ ¡éªŒè¾“å…¥æ•°æ®
    if not isinstance(data, dict) or len(data) < 2:
        raise ValueError("è¾“å…¥æ•°æ®å¿…é¡»ä¸ºè‡³å°‘åŒ…å«ä¸¤ä¸ªé”®çš„å­—å…¸")

    keys = list(data.keys())
    xdata = data[keys[0]]

    if not isinstance(xdata, np.ndarray):
        raise TypeError("xdata å¿…é¡»ä¸º numpy æ•°ç»„")

    # éå†æ¯ä¸€ä¸ªå› å˜é‡è¿›è¡Œå¤šé¡¹å¼æ‹Ÿåˆ
    for dependent_var in keys[1:]:
        ydata = data[dependent_var]

        if not isinstance(ydata, np.ndarray):
            results[dependent_var] = {"error": "ydata å¿…é¡»ä¸º numpy æ•°ç»„"}
            continue

        if len(xdata) != len(ydata):
            results[dependent_var] = {"error": "xdata ä¸ ydata é•¿åº¦ä¸ä¸€è‡´"}
            continue

        results[dependent_var] = {}

        # å°è¯•ä»1åˆ°max_degreeé˜¶çš„å¤šé¡¹å¼æ‹Ÿåˆ
        for degree in range(1, max_degree + 1):
            # åˆå§‹çŒœæµ‹å‚æ•°
            p0 = np.zeros(degree)  # æ›´åˆç†çš„åˆå§‹å€¼

            try:
                coefficients, _ = curve_fit(
                    polynomial_fit, xdata, ydata, maxfev=10000, p0=p0
                )
            except Exception:
                results[dependent_var] = {"error": "æ— æ³•æ‹Ÿåˆ"}
                break

            # è®¡ç®—æ‹Ÿåˆä¼˜åº¦æŒ‡æ ‡ RÂ² å’Œè°ƒæ•´ RÂ²
            y_fit = polynomial_fit(xdata, *coefficients)
            residuals = ydata - y_fit
            ss_res = np.sum(residuals**2)
            ss_tot = np.sum((ydata - np.mean(ydata)) ** 2)

            if ss_tot == 0:
                r_squared = 1.0  # æ‰€æœ‰å€¼ç›¸åŒï¼Œå®Œå…¨æ‹Ÿåˆ
            else:
                r_squared = 1 - (ss_res / ss_tot)

            n = len(ydata)
            if n - degree - 1 == 0:
                adj_r_squared = np.nan  # æ— æ³•è®¡ç®—è°ƒæ•´ RÂ²
            else:
                adj_r_squared = 1 - (1 - r_squared) * (n - 1) / (n - degree - 1)

            # å¦‚æœå½“å‰é˜¶æ•°æ»¡è¶³æ‹Ÿåˆç²¾åº¦è¦æ±‚ï¼Œåˆ™è®°å½•ç»“æœå¹¶è·³å‡ºå¾ªç¯
            if r_squared >= threshold and adj_r_squared >= threshold:
                results[dependent_var] = {
                    "degree": degree,
                    "coefficients": coefficients,
                    "r_squared": r_squared,
                    "adj_r_squared": adj_r_squared,
                }
                break
            # å¦‚æœè¾¾åˆ°æœ€å¤§é˜¶æ•°ä»æœªæ»¡è¶³æ¡ä»¶ï¼Œåˆ™æ ‡è®°ä¸ºè¶…é™
            elif degree == max_degree:
                results[dependent_var] = {"error": "å¤šé¡¹å¼ç³»æ•°è¶…é™"}
                break

    return results


def generate_polynomial_expression(
    xname: str, yname: str, coefficients: np.ndarray
) -> str:
    """ç”Ÿæˆå¤šé¡¹å¼è¡¨è¾¾å¼çš„å­—ç¬¦ä¸²è¡¨ç¤ºå½¢å¼

    æ ¹æ®ç»™å®šçš„ç³»æ•°æ•°ç»„ç”Ÿæˆå½¢å¦‚ "y = a0 + a1*x + a2*x^2 + ..." çš„å¤šé¡¹å¼è¡¨è¾¾å¼å­—ç¬¦ä¸²ï¼Œ
    ä¼šè‡ªåŠ¨å¤„ç†ç³»æ•°ä¸º0ã€1ã€-1ç­‰ç‰¹æ®Šæƒ…å†µï¼Œå¹¶æ ¼å¼åŒ–ç³»æ•°æ˜¾ç¤ºã€‚

    Parameters
    ----------
    xname : str
        è‡ªå˜é‡çš„åç§°ï¼Œç”¨äºæ„å»ºå¤šé¡¹å¼ä¸­çš„xé¡¹
    yname : str
        å› å˜é‡çš„åç§°ï¼Œç”¨äºç­‰å¼å·¦ä¾§çš„å˜é‡å
    coefficients : np.ndarray
        å¤šé¡¹å¼çš„ç³»æ•°æ•°ç»„ï¼Œç´¢å¼•å¯¹åº”å¹‚æ¬¡ï¼Œcoefficients[i]ä¸ºx^ié¡¹çš„ç³»æ•°

    Returns
    -------
    str
        æ ¼å¼åŒ–åçš„å¤šé¡¹å¼è¡¨è¾¾å¼å­—ç¬¦ä¸²
    """
    if len(coefficients) == 0:
        return f"{yname} = 0"

    expression = yname + " = "
    terms = []

    for i, coef in enumerate(coefficients):
        if coef == 0:
            continue

        term = ""
        sign = "+" if coef > 0 and i > 0 else ""  # é¦–é¡¹ä¸åŠ æ­£å·

        if i == 0:
            term = f"{sign}{coef:.6g}"
        elif i == 1:
            if coef == 1:
                term = f"{sign}{xname}"
            elif coef == -1:
                term = f"-{xname}"
            else:
                term = f"{sign}{coef:.6g}\\times {xname}"
        else:
            if coef == 1:
                term = f"{sign}{xname}^{i}"
            elif coef == -1:
                term = f"-{xname}^{i}"
            else:
                term = f"{sign}{coef:.6g}\\times {xname}^{i}"

        terms.append(term)

    full_expr = "".join(terms)
    # ç§»é™¤é¦–é¡¹å¤šä½™çš„æ­£å·
    result = full_expr.lstrip("+")
    return expression + result if result else expression + "0"


def run(
    file_path: str,
) -> tuple[dict[str, np.ndarray], dict[str, dict[str, int | float | np.ndarray]]]:
    """
    æ‰§è¡Œæ•°æ®è¯»å–å’Œå¤šé¡¹å¼æ‹Ÿåˆæµç¨‹ï¼Œå¹¶å°†ç»“æœä»¥è¡¨æ ¼å½¢å¼è¾“å‡ºåˆ°æ§åˆ¶å°ã€‚

    è¯¥å‡½æ•°é¦–å…ˆä»æŒ‡å®šè·¯å¾„è¯»å–Excelæ•°æ®ï¼Œç„¶åå¯¹æ•°æ®è¿›è¡Œå¤šé¡¹å¼æ‹Ÿåˆã€‚æ‹Ÿåˆç»“æœä¼šæ ¹æ®å˜é‡ååˆ†ç»„æ˜¾ç¤ºï¼Œ
    åŒ…æ‹¬æ‹ŸåˆçŠ¶æ€ï¼ˆå¦‚ç³»æ•°è¶…é™æˆ–æ— æ³•æ‹Ÿåˆï¼‰ä»¥åŠè¯¦ç»†çš„æ‹Ÿåˆä¿¡æ¯ï¼Œä¾‹å¦‚LaTeXè¡¨è¾¾å¼ã€StarCCM+ç³»æ•°ã€
    Fluentç³»æ•°ä»¥åŠå…¶ä»–ç»Ÿè®¡æŒ‡æ ‡ã€‚

    Parameters
    ----------
    file_path : str
        Excelæ–‡ä»¶çš„è·¯å¾„ï¼Œç”¨äºè¯»å–å¾…æ‹Ÿåˆçš„æ•°æ®ã€‚

    Returns
    -------
    tuple[dict[str, np.ndarray], dict[str, dict[str, int | float | np.ndarray]]]
        ä¸€ä¸ªå…ƒç»„ï¼ŒåŒ…å«ä¸¤ä¸ªå…ƒç´ ï¼š
        - ç¬¬ä¸€ä¸ªå…ƒç´ æ˜¯åŸå§‹æ•°æ®å­—å…¸ï¼Œé”®ä¸ºå˜é‡åï¼Œå€¼ä¸ºå¯¹åº”çš„æ•°å€¼æ•°ç»„ï¼›
        - ç¬¬äºŒä¸ªå…ƒç´ æ˜¯æ‹Ÿåˆç»“æœå­—å…¸ï¼Œå¤–å±‚é”®ä¸ºå˜é‡åå…ƒç»„ï¼Œå†…å±‚é”®ä¸ºæ‹Ÿåˆå±æ€§ï¼ˆå¦‚ç³»æ•°ã€RÂ²ç­‰ï¼‰ï¼Œå€¼ä¸ºç›¸åº”çš„æ‹Ÿåˆç»“æœæ•°æ®ï¼›
    """

    try:
        # è¯»å–Excelä¸­çš„æ•°æ®
        data = read_excel_data(file_path)
    except Exception as e:
        console.print(f"[red]è¯»å–Excelæ•°æ®å¤±è´¥: {e}[/red]")
        raise

    try:
        # å¯¹è¯»å–çš„æ•°æ®æ‰§è¡Œå¤šé¡¹å¼æ‹Ÿåˆæ“ä½œ
        results = fit_polynomial(data)
    except Exception as e:
        console.print(f"[red]å¤šé¡¹å¼æ‹Ÿåˆå¤±è´¥: {e}[/red]")
        raise

    # éå†æ‹Ÿåˆç»“æœå¹¶æ ¼å¼åŒ–è¾“å‡ºè‡³æ§åˆ¶å°
    for key1, value1 in results.items():
        var_name = " ".join(key1)
        console.print(Panel(f"[bold blue]{var_name}[/bold blue]", expand=False))

        # å¤„ç†æ‹Ÿåˆå¤±è´¥çš„æƒ…å†µ
        if value1 == "å¤šé¡¹å¼ç³»æ•°è¶…é™":
            console.print("[red]å¤šé¡¹å¼ç³»æ•°è¶…é™[/red]\n")
            continue
        elif value1 == "æ— æ³•æ‹Ÿåˆ":
            console.print("[red]æ— æ³•æ‹Ÿåˆ[/red]\n")
            continue

        # åˆ›å»ºè¡¨æ ¼ç”¨äºå±•ç¤ºå½“å‰å˜é‡çš„æ‹Ÿåˆç»“æœ
        table = Table(show_header=False, box=None)
        table.add_column("Property", style="cyan")
        table.add_column("Value", style="magenta")

        # å¡«å……è¡¨æ ¼å†…å®¹
        for key2, value2 in value1.items():
            if key2 == "coefficients":
                # ç”ŸæˆLaTeXæ ¼å¼çš„å¤šé¡¹å¼è¡¨è¾¾å¼
                latex_polynomial_expression = generate_polynomial_expression(
                    "T", "y", value2
                )
                table.add_row("LaTeX å…¬å¼", latex_polynomial_expression)

                # ç¼“å­˜æ ¼å¼åŒ–åçš„ç³»æ•°å­—ç¬¦ä¸²
                formatted_coeffs = [f"{x:.6g}" for x in value2]

                # æ·»åŠ StarCCM+æ ¼å¼çš„ç³»æ•°è¡¨ç¤º
                table.add_row("StarCCM+ ç³»æ•°", f"[{', '.join(formatted_coeffs)}]")

                # æ·»åŠ Fluentæ ¼å¼çš„ç³»æ•°è¡¨ç¤º
                degree = value1.get("degree", len(value2) - 1)
                table.add_row(
                    "Fluent ç³»æ•°",
                    f"{degree} {' '.join(formatted_coeffs)}",
                )
            else:
                # æ˜¾ç¤ºå…¶ä»–æ‹Ÿåˆå±æ€§ï¼ˆå¦‚RÂ²ã€å‡æ–¹è¯¯å·®ç­‰ï¼‰
                table.add_row(
                    key2, f"{value2:.6f}" if isinstance(value2, float) else str(value2)
                )

        # è¾“å‡ºå½“å‰å˜é‡çš„ç»“æœè¡¨æ ¼
        console.print(table)
        console.print()

    # è¿”å›åŸå§‹æ•°æ®å’Œæ‹Ÿåˆç»“æœ
    return data, results


def cli_main():
    """_summary_"""
    # è·å–å½“å‰è„šæœ¬æ‰€åœ¨ç›®å½•
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # æ„å»ºé»˜è®¤æ–‡ä»¶è·¯å¾„
    default_file_path = os.path.join(current_dir, "data.xlsx")

    # å¯¼å…¥ rich_argparse å¹¶è®¾ç½®ä¸ºé»˜è®¤æ ¼å¼åŒ–ç±»
    rich_argparse.RichHelpFormatter.styles["argparse.prog"] = "bold cyan"
    rich_argparse.RichHelpFormatter.styles["argparse.args"] = "yellow"
    rich_argparse.RichHelpFormatter.styles["argparse.metavar"] = "yellow"

    # åˆ›å»ºå‘½ä»¤è¡Œå‚æ•°è§£æå™¨
    parser = argparse.ArgumentParser(
        description="[bold green]å¤šé¡¹å¼æ‹Ÿåˆå·¥å…·[/bold green]",
        formatter_class=rich_argparse.RichHelpFormatter,
    )
    parser.add_argument(
        "file_path",
        type=str,
        nargs="?",
        default=default_file_path,
        help="Excelæ–‡ä»¶è·¯å¾„ (é»˜è®¤: data.xlsx)",
    )
    parser.add_argument("--no-plot", "-n", action="store_true", help="ä¸æ˜¾ç¤ºå›¾è¡¨")
    parser.add_argument(
        "--columns", "-c", type=int, default=2, help="å›¾è¡¨æ˜¾ç¤ºåˆ—æ•° (é»˜è®¤: 2)"
    )

    # è§£æå‘½ä»¤è¡Œå‚æ•°
    args = parser.parse_args()

    # æ˜¾ç¤ºç¨‹åºæ ‡é¢˜å’Œè¾“å…¥ä¿¡æ¯
    console.print(Panel("[bold blue]å¤šé¡¹å¼æ‹Ÿåˆå·¥å…·[/bold blue]", expand=False))
    console.print(f"è¾“å…¥æ–‡ä»¶: [cyan]{args.file_path}[/cyan]")

    # è¿è¡Œä¸»ç¨‹åº
    try:
        data, results = run(args.file_path)
    except FileNotFoundError:
        console.print(f"[red]é”™è¯¯ï¼šæ‰¾ä¸åˆ°æ–‡ä»¶ '{args.file_path}'[/red]")
        exit(1)
    except Exception as e:
        console.print(f"[red]é”™è¯¯ï¼š{str(e)}[/red]")
        exit(1)

    # å¦‚æœæ²¡æœ‰æŒ‡å®š--no-plotï¼Œåˆ™æ˜¾ç¤ºå›¾è¡¨
    if not args.no_plot:
        plot_fits_pygal_combined(data, results, filename=None, columns=args.columns)
    else:
        console.print("[yellow]å·²è·³è¿‡å›¾è¡¨æ˜¾ç¤º[/yellow]")


def input_main():
    """_summary_"""
    try:
        # æ˜¾ç¤ºç¨‹åºæ ‡é¢˜
        console = Console(width=60)
        console.print(f"\n[bold green]{script_name}[/bold green]\n", justify="center")

        # äº¤äº’å¼è¾“å…¥å‚æ•°
        while True:
            try:
                console.print("[bold cyan]å‚æ•°é…ç½®[/bold cyan]", justify="center")
                console.print("=" * 60)
                # è·å–å½“å‰è„šæœ¬æ‰€åœ¨ç›®å½•
                current_dir = os.getcwd()
                # æ„å»ºé»˜è®¤æ–‡ä»¶è·¯å¾„
                default_file_path = os.path.join(current_dir, "data.xlsx")
                # è¾“å…¥æ–‡ä»¶è·¯å¾„
                input_file_path = Prompt.ask(
                    "[bold blue]Excelæ–‡ä»¶è·¯å¾„[/bold blue]",
                    default=default_file_path,
                    console=console,
                )
                console.print(f"æ–‡ä»¶è·¯å¾„: [cyan]{input_file_path}[/cyan]\n")

                # æ˜¯å¦æ˜¾ç¤ºå›¾åƒ
                plot_show_input = Prompt.ask(
                    "[bold blue]æ˜¯å¦æ˜¾ç¤ºå›¾åƒ?[/bold blue] [dim](y/n)[/dim]",
                    choices=["y", "n"],
                    default="y",
                    console=console,
                )
                plot_show = plot_show_input.lower() in ["y", "yes"]
                status = "[green]å¯ç”¨[/green]" if plot_show else "[red]ç¦ç”¨[/red]"
                console.print(f"å›¾åƒæ˜¾ç¤º: {status}\n")

                # å›¾è¡¨åˆ—æ•°
                columns = IntPrompt.ask(
                    "[bold blue]å›¾è¡¨æ˜¾ç¤ºåˆ—æ•°[/bold blue]", default=2, console=console
                )
                console.print(f"æ˜¾ç¤ºåˆ—æ•°: [magenta]{columns}[/magenta]\n")

                # æ˜¾ç¤ºåˆ†éš”çº¿
                console.print("=" * 60)
                console.print("[bold green]å¼€å§‹å¤„ç†æ•°æ®...[/bold green]\n")

            except ValueError as e:
                console.print(f"[red]è¾“å…¥æ ¼å¼é”™è¯¯: {str(e)}ï¼Œè¯·é‡æ–°è¾“å…¥[/red]\n")
                continue
            except KeyboardInterrupt:
                console.print("\n[yellow]ç”¨æˆ·å–æ¶ˆæ“ä½œ[/yellow]")
                return

            # è¿è¡Œä¸»ç¨‹åº
            try:
                data, results = run(input_file_path)

                # æ˜¾ç¤ºå›¾è¡¨
                if plot_show:
                    plot_fits_pygal_combined(
                        data, results, filename=None, columns=int(columns)
                    )
                    console.print("[green]å›¾è¡¨å·²ç”Ÿæˆå¹¶æ˜¾ç¤º[/green]\n")
                else:
                    console.print("[yellow]å·²è·³è¿‡å›¾è¡¨æ˜¾ç¤º[/yellow]\n")

            except Exception as e:
                console.print(f"[red]å¤„ç†è¿‡ç¨‹ä¸­å‘ç”Ÿé”™è¯¯: {str(e)}[/red]\n")
                continue

            # æ˜¾ç¤ºå®Œæˆä¿¡æ¯
            console.print("=" * 60)
            console.print(
                "[bold green]æ‰€æœ‰ä»»åŠ¡å·²å®Œæˆ![/bold green]",
                justify="center",
            )

            # ç­‰å¾…ç”¨æˆ·é€€å‡º
            console.input("\n[bold cyan]æŒ‰ä»»æ„é”®é€€å‡º...[/bold cyan]")
            break

    except Exception as e:
        console.print(f"[red]ç¨‹åºå‘ç”Ÿå¼‚å¸¸: {str(e)}[/red]")
        console.input("\n[bold red]ç¨‹åºè¿è¡Œå‡ºé”™ï¼ŒæŒ‰ä»»æ„é”®é€€å‡º...[/bold red]")


def main():
    if len(sys.argv) > 1:
        cli_main()
    else:
        input_main()


if __name__ == "__main__":
    main()

# ğŸ§ª pyinstaller --onefile --add-data "src/assets;src/assets" --console -i "src/assets/icon.ico" --clean ./src/main.py
# ğŸ§ª pyinstaller --onefile --console --add-data "src/assets;assets" --collect-data pygal -i "src/assets/icon.ico" --clean --name "Polynomial-Fitter" ./src/main.py
